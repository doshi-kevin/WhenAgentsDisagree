"""Export service for CSV/JSON/Excel data export."""
import csv
import io
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.db.models import Debate, Turn, TurnMetrics, DebateAgent


HEADERS = [
    # Debate-level
    "debate_id", "scenario_id", "scenario_title", "category", "difficulty",
    "strategy", "final_answer", "ground_truth", "is_correct",
    "total_tokens", "total_latency_ms", "total_turns",
    "deadlock_detected", "deadlock_resolution",
    # Agent-level
    "agent_name", "provider", "model_id", "role", "assigned_position",
    "bias_role", "source_type", "source_reliability",
    # Turn-level
    "turn_number", "round_number", "turn_role",
    "content", "reasoning", "confidence", "position_held",
    "position_changed", "change_reason",
    "prompt_tokens", "completion_tokens", "total_tokens_turn", "latency_ms",
    # Metrics
    "aggressiveness_score", "sentiment_score", "persuasion_attempt_score",
    "citation_count", "citation_quality_score",
    "semantic_similarity_to_prev", "argument_novelty_score",
    "word_count", "hedging_language_count",
]


async def _load_debates(db: AsyncSession, experiment_id: str = None) -> list[Debate]:
    q = select(Debate).options(
        selectinload(Debate.agents),
        selectinload(Debate.turns).selectinload(Turn.metrics),
        selectinload(Debate.turns).selectinload(Turn.agent),
        selectinload(Debate.scenario),
    )
    if experiment_id:
        q = q.where(Debate.experiment_id == experiment_id)
    q = q.where(Debate.status == "completed")
    result = await db.execute(q)
    return list(result.scalars().unique().all())


def _safe(val, fallback=""):
    """Safely return a value or fallback if None."""
    return val if val is not None else fallback


def _build_rows(debates: list[Debate]) -> list[list]:
    """Build flat row data from debates for CSV/Excel export."""
    rows = []
    for debate in debates:
        scenario = debate.scenario
        # Build agent briefing lookup
        briefing_lookup = {}
        if scenario and scenario.agent_briefings:
            for b in scenario.agent_briefings:
                briefing_lookup[b.get("position", "")] = b

        for turn in debate.turns:
            agent = turn.agent
            metrics = turn.metrics

            # Safely get agent fields (agent can be None for judge turns)
            a_name = agent.agent_name if agent else ""
            a_provider = agent.provider if agent else ""
            a_model = agent.model_id if agent else ""
            a_role = agent.role if agent else ""
            a_position = agent.assigned_position if agent else ""
            a_bias = (agent.bias_role if agent and agent.bias_role else "")

            # Source metadata from scenario briefing
            agent_briefing = briefing_lookup.get(a_position, {}) if a_position else {}

            content = _safe(turn.content, "")
            if len(content) > 1000:
                content = content[:1000]
            reasoning = _safe(turn.reasoning, "")
            if len(reasoning) > 500:
                reasoning = reasoning[:500]

            rows.append([
                debate.id,
                scenario.id if scenario else "",
                scenario.title if scenario else "",
                scenario.category if scenario else "",
                scenario.difficulty if scenario else "",
                debate.strategy,
                _safe(debate.final_answer),
                scenario.ground_truth if scenario else "",
                debate.is_correct,
                _safe(debate.total_tokens, 0),
                _safe(debate.total_latency_ms, 0),
                _safe(debate.total_turns, 0),
                debate.deadlock_detected,
                _safe(debate.deadlock_resolution, ""),
                # Agent
                a_name, a_provider, a_model, a_role, a_position, a_bias,
                agent_briefing.get("source_type", ""),
                _safe(agent_briefing.get("source_reliability"), ""),
                # Turn
                turn.turn_number,
                turn.round_number,
                _safe(turn.role, ""),
                content,
                reasoning,
                _safe(turn.confidence_score),
                _safe(turn.position_held, ""),
                turn.position_changed,
                _safe(turn.change_reason, ""),
                _safe(turn.prompt_tokens, 0),
                _safe(turn.completion_tokens, 0),
                _safe(turn.total_tokens, 0),
                _safe(turn.latency_ms, 0),
                # Metrics
                _safe(metrics.aggressiveness_score if metrics else None, ""),
                _safe(metrics.sentiment_score if metrics else None, ""),
                _safe(metrics.persuasion_attempt_score if metrics else None, ""),
                _safe(metrics.citation_count if metrics else None, ""),
                _safe(metrics.citation_quality_score if metrics else None, ""),
                _safe(metrics.semantic_similarity_to_prev if metrics else None, ""),
                _safe(metrics.argument_novelty_score if metrics else None, ""),
                _safe(metrics.word_count if metrics else None, ""),
                _safe(metrics.hedging_language_count if metrics else None, ""),
            ])
    return rows


async def export_debates_csv(db: AsyncSession, experiment_id: str = None) -> str:
    """Export debate data as research-ready CSV."""
    debates = await _load_debates(db, experiment_id)
    rows = _build_rows(debates)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(row)

    return output.getvalue()


async def export_debates_excel(db: AsyncSession, experiment_id: str = None) -> bytes:
    """Export debate data as Excel (.xlsx) with formatted sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    debates = await _load_debates(db, experiment_id)
    rows = _build_rows(debates)

    wb = Workbook()

    # --- Sheet 1: All Turns (raw data) ---
    ws = wb.active
    ws.title = "Debate Turns"

    # Header styling
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_text = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-width columns (capped at 40)
    for col_idx in range(1, len(HEADERS) + 1):
        max_len = len(str(HEADERS[col_idx - 1]))
        for row_idx in range(2, min(len(rows) + 2, 50)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    # Freeze header row
    ws.freeze_panes = "A2"
    # Add autofilter
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADERS)).column_letter}{len(rows) + 1}"

    # --- Sheet 2: Debate Summary ---
    ws2 = wb.create_sheet("Debate Summary")
    summary_headers = [
        "debate_id", "scenario_title", "category", "strategy",
        "final_answer", "ground_truth", "is_correct",
        "total_turns", "total_tokens", "total_latency_ms",
        "deadlock_detected", "num_agents", "agent_models",
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    correct_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    incorrect_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    for row_idx, debate in enumerate(debates, 2):
        scenario = debate.scenario
        models = ", ".join(sorted(set(
            f"{a.provider}:{a.model_id}" for a in debate.agents
        )))
        row_data = [
            debate.id,
            scenario.title if scenario else "",
            scenario.category if scenario else "",
            debate.strategy,
            _safe(debate.final_answer),
            scenario.ground_truth if scenario else "",
            debate.is_correct,
            _safe(debate.total_turns, 0),
            _safe(debate.total_tokens, 0),
            _safe(debate.total_latency_ms, 0),
            debate.deadlock_detected,
            len(debate.agents),
            models,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
        # Color correct/incorrect rows
        fill = correct_fill if debate.is_correct else incorrect_fill
        for col_idx in range(1, len(summary_headers) + 1):
            ws2.cell(row=row_idx, column=col_idx).fill = fill

    ws2.freeze_panes = "A2"
    for col_idx in range(1, len(summary_headers) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 18

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def export_debates_json(db: AsyncSession, experiment_id: str = None) -> list:
    """Export debate data as research-ready JSON with full metadata."""
    debates = await _load_debates(db, experiment_id)

    export = []
    for debate in debates:
        scenario = debate.scenario

        briefing_lookup = {}
        if scenario and scenario.agent_briefings:
            for b in scenario.agent_briefings:
                briefing_lookup[b.get("position", "")] = b

        agents_data = []
        for agent in debate.agents:
            pos = agent.assigned_position or ""
            ab = briefing_lookup.get(pos, {})
            agents_data.append({
                "agent_name": agent.agent_name,
                "provider": agent.provider,
                "model_id": agent.model_id,
                "role": agent.role,
                "assigned_position": pos,
                "bias_role": agent.bias_role,
                "source_type": ab.get("source_type"),
                "source_reliability": ab.get("source_reliability"),
            })

        turns_data = []
        for turn in debate.turns:
            agent = turn.agent
            metrics = turn.metrics
            turns_data.append({
                "turn_number": turn.turn_number,
                "round_number": turn.round_number,
                "agent_name": agent.agent_name if agent else "",
                "provider": agent.provider if agent else "",
                "model_id": agent.model_id if agent else "",
                "role": turn.role,
                "content": turn.content,
                "reasoning": turn.reasoning,
                "confidence": turn.confidence_score,
                "position_held": turn.position_held,
                "position_changed": turn.position_changed,
                "change_reason": turn.change_reason,
                "prompt_tokens": turn.prompt_tokens,
                "completion_tokens": turn.completion_tokens,
                "total_tokens": turn.total_tokens,
                "latency_ms": turn.latency_ms,
                "metrics": {
                    "aggressiveness_score": metrics.aggressiveness_score,
                    "sentiment_score": metrics.sentiment_score,
                    "persuasion_attempt_score": metrics.persuasion_attempt_score,
                    "citation_count": metrics.citation_count,
                    "citation_quality_score": metrics.citation_quality_score,
                    "semantic_similarity_to_prev": metrics.semantic_similarity_to_prev,
                    "argument_novelty_score": metrics.argument_novelty_score,
                    "word_count": metrics.word_count,
                    "hedging_language_count": metrics.hedging_language_count,
                } if metrics else None,
            })

        export.append({
            "debate_id": debate.id,
            "scenario": {
                "id": scenario.id if scenario else None,
                "title": scenario.title if scenario else None,
                "category": scenario.category if scenario else None,
                "question": scenario.question if scenario else None,
                "ground_truth": scenario.ground_truth if scenario else None,
                "difficulty": scenario.difficulty if scenario else None,
            },
            "strategy": debate.strategy,
            "final_answer": debate.final_answer,
            "is_correct": debate.is_correct,
            "total_tokens": debate.total_tokens,
            "total_latency_ms": debate.total_latency_ms,
            "total_turns": debate.total_turns,
            "deadlock_detected": debate.deadlock_detected,
            "deadlock_resolution": debate.deadlock_resolution,
            "started_at": debate.started_at.isoformat() if debate.started_at else None,
            "completed_at": debate.completed_at.isoformat() if debate.completed_at else None,
            "agents": agents_data,
            "turns": turns_data,
        })

    return export
